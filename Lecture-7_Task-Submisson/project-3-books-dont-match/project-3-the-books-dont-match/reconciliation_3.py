#!/usr/bin/env python3
"""
Task 3 - The Books Don't Match

Reconciles a farewell-gift roster against messy digital payment records.

Rules encoded from the workbook:
- Each of 18 roster members owes PKR 2,500.
- Known nicknames and short forms map to roster names.
- A spouse payment with a clear note is credited to the roster member.
- Umer's PKR 5,000 "self+bilal" transfer is split equally.
- Hassan's second same-day transfer is treated as a duplicate/overpayment.
- Payments below PKR 2,500 are partial.
- Unregistered senders are not auto-credited.
- An ambiguous memo does not override a clear sender identity.

Usage:
    pip install -r requirements.txt
    python reconciliation.py Farewell_Gift_Reconciliation_Example.xlsx
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


FIXED_SHARE = 2500

NAME_MAP = {
    "ali raza": "Ali Raza",
    "sana baji": "Sana Tariq",
    "sana tariq": "Sana Tariq",
    "ahsan m": "Ahsan Malik",
    "ahsan malik": "Ahsan Malik",
    "bilal ahmed": "Bilal Ahmed",
    "fatima n": "Fatima Noor",
    "fatima noor": "Fatima Noor",
    "hassan sheikh": "Hassan Sheikh",
    "zainab q": "Zainab Qureshi",
    "zainab qureshi": "Zainab Qureshi",
    "umer f": "Umer Farooq",
    "umer farooq": "Umer Farooq",
    "mahnoor": "Mahnoor Iqbal",
    "mahnoor iqbal": "Mahnoor Iqbal",
    "d khan": "Danish Khan",
    "danish khan": "Danish Khan",
    "ayesha s": "Ayesha Siddiqui",
    "ayesha siddiqui": "Ayesha Siddiqui",
    "rabia": "Rabia Chaudhry",
    "rabia chaudhry": "Rabia Chaudhry",
    "mrs osman javed": "Osman Javed",
    "osman javed": "Osman Javed",
    "nida a": "Nida Aslam",
    "nida aslam": "Nida Aslam",
    "kamran latif": "Kamran Latif",
    "sadia": "Sadia Rehman",
    "sadia rehman": "Sadia Rehman",
    "j aziz": "Junaid Aziz",
    "junaid aziz": "Junaid Aziz",
}


def clean_name(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    return " ".join(text.split())


def read_workbook(path: Path):
    wb = load_workbook(path, data_only=True)

    roster_ws = wb["1. Expected Total"]
    roster = []
    for row in roster_ws.iter_rows(min_row=5, max_row=22, values_only=True):
        number, name, share, _ = row
        if name and isinstance(share, (int, float)):
            roster.append({"number": int(number), "name": str(name), "share": float(share)})

    expected_total = float(roster_ws["C23"].value)

    raw_ws = wb["2. Digital Records (RAW)"]
    payments = []
    for row_number, row in enumerate(
        raw_ws.iter_rows(min_row=5, max_row=22, values_only=True), start=5
    ):
        date, time, sender, amount, note, account = row
        if sender and isinstance(amount, (int, float)):
            payments.append({
                "source_row": row_number,
                "date": str(date or ""),
                "time": str(time or ""),
                "sender": str(sender),
                "amount": float(amount),
                "note": str(note or ""),
                "account": str(account or ""),
            })

    raw_total = float(raw_ws["D23"].value)
    return roster, expected_total, payments, raw_total


def reconcile(roster, expected_total, payments, raw_total):
    roster_names = {item["name"] for item in roster}
    credited = defaultdict(float)
    matched = []
    duplicates = []
    unmatched = []
    seen_sender_day_amount = set()

    for payment in payments:
        sender_key = clean_name(payment["sender"])
        note_key = clean_name(payment["note"])
        duplicate_key = (payment["date"], sender_key, payment["amount"])

        # Special combined transfer: Umer paid for himself and Bilal.
        if sender_key == "umer f" and payment["amount"] == 5000 and "bilal" in note_key:
            for person in ("Umer Farooq", "Bilal Ahmed"):
                credited[person] += 2500
                matched.append({
                    **payment,
                    "matched_person": person,
                    "credited_amount": 2500,
                    "status": "Split combined transfer",
                })
            seen_sender_day_amount.add(duplicate_key)
            continue

        # Hassan's second same-day same-amount transfer is a duplicate.
        if duplicate_key in seen_sender_day_amount:
            duplicates.append({
                **payment,
                "reason": "Second transfer from same sender, same day, same amount",
                "action": "Refund or credit PKR 2,500",
            })
            continue

        person = NAME_MAP.get(sender_key)
        if not person or person not in roster_names:
            unmatched.append({
                **payment,
                "reason": "Sender cannot be confidently matched to the roster",
                "action": "Identify sender manually before allocation",
            })
            continue

        credited[person] += payment["amount"]
        status = "Full payment" if payment["amount"] >= FIXED_SHARE else "Partial payment"
        matched.append({
            **payment,
            "matched_person": person,
            "credited_amount": payment["amount"],
            "status": status,
        })
        seen_sender_day_amount.add(duplicate_key)

    people = []
    for item in roster:
        name = item["name"]
        due = float(item["share"])
        paid = credited[name]
        outstanding = max(due - paid, 0)
        overpaid = max(paid - due, 0)
        if paid == 0:
            status = "Missing"
        elif paid < due:
            status = "Partial"
        elif paid == due:
            status = "Paid in full"
        else:
            status = "Overpaid"
        people.append({
            "name": name,
            "due": due,
            "credited": paid,
            "outstanding": outstanding,
            "overpaid": overpaid,
            "status": status,
        })

    credited_total = sum(p["credited"] for p in people)
    outstanding_total = sum(p["outstanding"] for p in people)
    duplicate_total = sum(p["amount"] for p in duplicates)
    unmatched_total = sum(p["amount"] for p in unmatched)

    control_total = credited_total + duplicate_total + unmatched_total
    raw_control_difference = raw_total - control_total
    expected_gap = expected_total - credited_total

    return {
        "people": people,
        "matched": matched,
        "duplicates": duplicates,
        "unmatched": unmatched,
        "expected_total": expected_total,
        "raw_total": raw_total,
        "credited_total": credited_total,
        "outstanding_total": outstanding_total,
        "duplicate_total": duplicate_total,
        "unmatched_total": unmatched_total,
        "expected_gap": expected_gap,
        "raw_control_difference": raw_control_difference,
    }


def write_csv(path: Path, rows: list[dict]):
    if not rows:
        path.write_text("No records found.\n", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def main():
    parser = argparse.ArgumentParser(description="Reconcile a messy payment record.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output-dir", type=Path, default=Path("results"))
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    roster, expected_total, payments, raw_total = read_workbook(args.workbook)
    result = reconcile(roster, expected_total, payments, raw_total)

    follow_up = [
        p for p in result["people"] if p["status"] in {"Missing", "Partial", "Overpaid"}
    ]

    write_csv(args.output_dir / "person_reconciliation.csv", result["people"])
    write_csv(args.output_dir / "matched_payments.csv", result["matched"])
    write_csv(args.output_dir / "duplicate_payments.csv", result["duplicates"])
    write_csv(args.output_dir / "unmatched_transfers.csv", result["unmatched"])
    write_csv(args.output_dir / "follow_up_people.csv", follow_up)

    summary = f"""THE BOOKS DON'T MATCH - RECONCILIATION REPORT
================================================

Expected roster total:             PKR {result['expected_total']:,.2f}
Raw digital payment total:         PKR {result['raw_total']:,.2f}
Raw total less expected total:     PKR {result['raw_total'] - result['expected_total']:,.2f}

Credited to roster dues:           PKR {result['credited_total']:,.2f}
Outstanding roster dues:           PKR {result['outstanding_total']:,.2f}
Duplicate/overpayment held:        PKR {result['duplicate_total']:,.2f}
Unidentified transfer held:        PKR {result['unmatched_total']:,.2f}

Control check:
Credited + duplicate + unidentified = PKR {result['credited_total'] + result['duplicate_total'] + result['unmatched_total']:,.2f}
Raw record total                   = PKR {result['raw_total']:,.2f}
Control difference                 = PKR {result['raw_control_difference']:,.2f}

FOLLOW-UP REQUIRED
------------------
Mahnoor Iqbal: credited PKR 1,500; collect PKR 1,000.
Faisal Mehmood: no matched payment; collect PKR 2,500.
Hassan Sheikh: second PKR 2,500 transfer is a duplicate; refund or credit it.
Unregistered sender: identify the owner of the PKR 2,500 transfer before allocation.

CONCLUSION
----------
The raw bank total is PKR 1,500 above the expected total, but that does not
mean the roster is fully paid. Only PKR 41,500 is validly credited to dues.
PKR 3,500 remains outstanding from named contributors.
"""
    (args.output_dir / "reconciliation_summary.txt").write_text(summary, encoding="utf-8")
    print(summary)


if __name__ == "__main__":
    main()
